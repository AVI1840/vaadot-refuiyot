"""
Extract data from Excel → JSON for the web app.
Project DD - Medical Committees Document Checklist System
"""
import openpyxl
import json
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'data',
    'קשר בין קבוצת אבחנות ומסמכים בכל התחומים33 (version 2).xlsx')
OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'app', 'data')

os.makedirs(OUT_DIR, exist_ok=True)
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)

# --- 1. Main diagnosis-document mapping (גיליון1) ---
ws = wb['גיליון1']
records = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    if row[0]:
        records.append({
            "diagnosisGroup": str(row[0]).strip(),
            "domain": str(row[1]).strip() if row[1] else "",
            "docGroup": str(row[2]).strip() if row[2] else "",
            "docId": row[3],
            "docName": str(row[4]).strip() if row[4] else ""
        })

# --- 2. AI-rated documents (Sheet1) ---
ws2 = wb['Sheet1']
ai_rated = []
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i == 0:
        continue
    if row[0]:
        ai_rated.append({
            "diagnosisGroup": str(row[0]).strip(),
            "domain": str(row[1]).strip() if row[1] else "",
            "docName": str(row[2]).strip() if row[2] else "",
            "docId": row[3],
            "aiConfidence": str(row[4]).strip() if row[4] else "",
            "docType": str(row[5]).strip() if row[5] else "",
            "implementationNotes": str(row[6]).strip() if row[6] else ""
        })
wb.close()

# --- 3. Build structured checklist data ---
# Group by diagnosis + domain
from collections import defaultdict

checklists = defaultdict(lambda: {"documents": [], "aiDocuments": []})

for r in records:
    key = f"{r['diagnosisGroup']}|{r['domain']}"
    checklists[key]["diagnosisGroup"] = r["diagnosisGroup"]
    checklists[key]["domain"] = r["domain"]
    checklists[key]["documents"].append({
        "docId": r["docId"],
        "docName": r["docName"],
        "docGroup": r["docGroup"]
    })

for r in ai_rated:
    key = f"{r['diagnosisGroup']}|{r['domain']}"
    checklists[key]["aiDocuments"].append({
        "docId": r["docId"],
        "docName": r["docName"],
        "aiConfidence": r["aiConfidence"],
        "docType": r["docType"],
        "implementationNotes": r["implementationNotes"]
    })

checklist_list = list(checklists.values())

# --- 4. Build stats ---
diagnosis_groups = sorted(set(r["diagnosisGroup"] for r in records))
domains = sorted(set(r["domain"] for r in records if r["domain"]))

domain_stats = defaultdict(int)
for r in records:
    domain_stats[r["domain"]] += 1

diag_stats = defaultdict(int)
for r in records:
    diag_stats[r["diagnosisGroup"]] += 1

top_diagnoses = sorted(diag_stats.items(), key=lambda x: -x[1])[:20]

mvp_diagnoses = ["Back pain", "Ischemic heart disease", "CTS", "DM", "COPD"]

stats = {
    "totalRecords": len(records),
    "totalAiRated": len(ai_rated),
    "diagnosisGroupCount": len(diagnosis_groups),
    "domainCount": len(domains),
    "domains": dict(domain_stats),
    "topDiagnoses": [{"name": d[0], "count": d[1]} for d in top_diagnoses],
    "mvpDiagnoses": mvp_diagnoses,
    "diagnosisGroups": diagnosis_groups
}

# --- 5. Write JSON files ---
with open(os.path.join(OUT_DIR, 'checklists.json'), 'w', encoding='utf-8') as f:
    json.dump(checklist_list, f, ensure_ascii=False, indent=2)

with open(os.path.join(OUT_DIR, 'stats.json'), 'w', encoding='utf-8') as f:
    json.dump(stats, f, ensure_ascii=False, indent=2)

with open(os.path.join(OUT_DIR, 'ai_rated.json'), 'w', encoding='utf-8') as f:
    json.dump(ai_rated, f, ensure_ascii=False, indent=2)

print(f"✅ Extracted {len(records)} records → checklists.json")
print(f"✅ Extracted {len(ai_rated)} AI-rated records → ai_rated.json")
print(f"✅ Stats → stats.json")
print(f"✅ {len(diagnosis_groups)} diagnosis groups, {len(domains)} domains")
